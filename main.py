from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import io
import os
import tempfile
from typing import Optional

# ============================================
# LAZY LOADING UTILITIES
# ============================================

def lazy_import_pdf2docx():
    """Lazy import for pdf2docx"""
    from pdf2docx import Converter
    return Converter

def lazy_import_pypdf2():
    """Lazy import for PyPDF2"""
    from PyPDF2 import PdfReader, PdfWriter
    return PdfReader, PdfWriter

def lazy_import_pil():
    """Lazy import for PIL"""
    from PIL import Image, ImageDraw, ImageFont
    return Image, ImageDraw, ImageFont

def lazy_import_qrcode():
    """Lazy import for qrcode"""
    import qrcode
    return qrcode

def lazy_import_pdf2image():
    """Lazy import for pdf2image"""
    from pdf2image import convert_from_path
    return convert_from_path

def lazy_import_docx():
    """Lazy import for python-docx"""
    from docx import Document
    return Document

def lazy_import_reportlab():
    """Lazy import for reportlab"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    return canvas, A4

def lazy_import_openpyxl():
    """Lazy import for openpyxl"""
    from openpyxl import load_workbook
    return load_workbook

def lazy_import_aspose_slides():
    """Lazy import for aspose.slides"""
    import aspose.slides as slides
    return slides

def lazy_import_fitz():
    """Lazy import for PyMuPDF (fitz)"""
    import fitz
    return fitz

def lazy_import_zipfile():
    """Lazy import for zipfile"""
    import zipfile
    return zipfile

def lazy_import_base64():
    """Lazy import for base64"""
    import base64
    return base64

def lazy_import_json():
    """Lazy import for json"""
    import json
    return json

# ============================================
# APP INITIALIZATION
# ============================================

app = FastAPI(title="TooConvert API", version="1.0.0")

UPLOAD_DIR = "temp"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ---------------- CORS ----------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://www.tooconvert.in",
        "https://tooconvert.in",
        "https://api.tooconvert.in"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================
# UTILITY FUNCTIONS
# ============================================

async def save_upload(file: UploadFile) -> str:
    """Save uploaded file to temp directory and return path"""
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    return file_path

# ============================================
# HEALTH CHECK
# ============================================

@app.get("/health")
def health():
    return {"message": "Server is running", "status": "healthy"}

# ============================================
# PDF OPERATIONS
# ============================================

@app.post("/pdf-to-docx/")
async def pdf_to_docx(file: UploadFile = File(...)):
    """Convert PDF to DOCX"""
    Converter = lazy_import_pdf2docx()
    
    file_path = f"temp/{file.filename}"
    with open(file_path, "wb") as f:
        f.write(await file.read())
    
    output_file = file_path.replace(".pdf", ".docx")
    cv = Converter(file_path)
    cv.convert(output_file, start=0, end=None)
    cv.close()
    
    return FileResponse(output_file, filename="converted.docx")

@app.post("/merge-pdf/")
async def merge_pdf(files: list[UploadFile] = File(...)):
    """Merge multiple PDFs into one"""
    PdfReader, PdfWriter = lazy_import_pypdf2()
    
    merger = PdfWriter()
    for file in files:
        reader = PdfReader(io.BytesIO(await file.read()))
        for page in reader.pages:
            merger.add_page(page)
    
    output_file = "temp/merged.pdf"
    with open(output_file, "wb") as f:
        merger.write(f)
    
    return FileResponse(output_file, media_type="application/pdf", filename="merged.pdf")

@app.post("/split-pdf/")
async def split_pdf(file: UploadFile = File(...), pages_per_split: int = Form(...)):
    """Split PDF into multiple files"""
    PdfReader, PdfWriter = lazy_import_pypdf2()
    zipfile = lazy_import_zipfile()
    
    reader = PdfReader(io.BytesIO(await file.read()))
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for i in range(0, len(reader.pages), pages_per_split):
            writer = PdfWriter()
            for j in range(i, min(i + pages_per_split, len(reader.pages))):
                writer.add_page(reader.pages[j])
            buf = io.BytesIO()
            writer.write(buf)
            buf.seek(0)
            zip_file.writestr(f"split_{i+1}.pdf", buf.read())

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=split_pdfs.zip"}
    )

@app.post("/extract-text/")
async def extract_text(file: UploadFile = File(...)):
    """Extract text from PDF"""
    PdfReader, _ = lazy_import_pypdf2()
    
    reader = PdfReader(io.BytesIO(await file.read()))
    text = "".join([page.extract_text() or "" for page in reader.pages])
    
    return JSONResponse({"text": text})

@app.post("/compress-pdf/")
async def compress_pdf(file: UploadFile = File(...), level: str = Form("medium")):
    """Compress PDF file"""
    fitz = lazy_import_fitz()
    
    try:
        pdf_bytes = await file.read()
        pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        compressed_pdf = io.BytesIO()

        # Map compression levels to image quality (0-100)
        quality_map = {
            "high": 20,
            "medium": 50,
            "low": 80
        }
        quality = quality_map.get(level.lower(), 50)

        # Save compressed PDF using PyMuPDF
        pdf_doc.save(
            compressed_pdf,
            garbage=4,       # remove unused objects
            deflate=True,    # compress streams
            clean=True       # clean up
        )
        compressed_pdf.seek(0)
        pdf_doc.close()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF compression failed: {e}")

    return StreamingResponse(
        compressed_pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=compressed_{file.filename}"}
    )

# ============================================
# IMAGE OPERATIONS
# ============================================

@app.post("/resize-image/")
async def resize_image(file: UploadFile = File(...), width: int = Form(...), height: int = Form(...)):
    """Resize image to specified dimensions"""
    Image, _, _ = lazy_import_pil()
    
    image = Image.open(io.BytesIO(await file.read()))
    resized = image.resize((width, height))
    buf = io.BytesIO()
    resized.save(buf, format="JPEG")
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type="image/jpeg", 
        headers={"Content-Disposition": "attachment; filename=resized.jpg"}
    )

@app.post("/convert-format/")
async def convert_format(file: UploadFile = File(...), format: str = Form(...)):
    """Convert image to different format"""
    Image, _, _ = lazy_import_pil()
    
    image = Image.open(io.BytesIO(await file.read()))
    buf = io.BytesIO()
    image.save(buf, format=format)
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type=f"image/{format.lower()}", 
        headers={"Content-Disposition": f"attachment; filename=converted.{format.lower()}"}
    )

@app.post("/watermark/")
async def add_watermark(
    file: UploadFile = File(...), 
    text: str = Form(...), 
    opacity: int = Form(...), 
    font_size: int = Form(...)
):
    """Add watermark to image"""
    Image, ImageDraw, ImageFont = lazy_import_pil()
    
    image = Image.open(io.BytesIO(await file.read()))
    if image.mode in ("RGBA", "LA"):
        image = image.convert("RGB")
    
    watermark = Image.new("RGBA", image.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(watermark)
    
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except:
        font = ImageFont.load_default()
    
    step_x = font_size * 10
    step_y = font_size * 8
    for y in range(0, image.height, step_y):
        for x in range(0, image.width, step_x):
            draw.text((x, y), text, fill=(255, 255, 255, opacity), font=font)
    
    watermarked = Image.alpha_composite(image.convert("RGBA"), watermark)
    buf = io.BytesIO()
    watermarked.convert("RGB").save(buf, format="JPEG")
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type="image/jpeg", 
        headers={"Content-Disposition": "attachment; filename=watermarked.jpg"}
    )

@app.post("/compress-image/")
async def compress_image(file: UploadFile = File(...), target_size: int = Form(...)):
    """Compress image to target size in KB"""
    Image, _, _ = lazy_import_pil()
    
    image = Image.open(io.BytesIO(await file.read()))
    if image.mode in ("RGBA", "LA"):
        image = image.convert("RGB")
    
    quality = 95
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=quality)
    compressed_data = buf.getvalue()
    compressed_size = len(compressed_data) / 1024
    
    while compressed_size > target_size and quality > 10:
        quality -= 5
        buf = io.BytesIO()
        image.save(buf, format="JPEG", quality=quality)
        compressed_data = buf.getvalue()
        compressed_size = len(compressed_data) / 1024
    
    buf = io.BytesIO(compressed_data)
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type="image/jpeg", 
        headers={"Content-Disposition": "attachment; filename=compressed.jpg"}
    )

# ============================================
# QR CODE GENERATION
# ============================================

@app.post("/generate-qr/")
async def generate_qr(text: str = Form(...)):
    """Generate QR code from text"""
    qrcode = lazy_import_qrcode()
    
    qr_img = qrcode.make(text)
    buf = io.BytesIO()
    qr_img.save(buf, format="PNG")
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type="image/png", 
        headers={"Content-Disposition": "attachment; filename=qrcode.png"}
    )

# ============================================
# BASE64 OPERATIONS
# ============================================

@app.post("/base64-encode/")
async def base64_encode(file: UploadFile = File(None), text: str = Form(None)):
    """Encode file or text to base64"""
    base64 = lazy_import_base64()
    
    if not file and not text:
        return {"error": "No input provided"}
    
    data = await file.read() if file else text.encode()
    return {"base64": base64.b64encode(data).decode()}

@app.post("/base64-decode/")
async def base64_decode(encoded: str = Form(...)):
    """Decode base64 to file"""
    base64 = lazy_import_base64()
    
    data = base64.b64decode(encoded)
    buf = io.BytesIO(data)
    buf.seek(0)
    
    return StreamingResponse(
        buf, 
        media_type="application/octet-stream", 
        headers={"Content-Disposition": "attachment; filename=decoded.bin"}
    )

# ============================================
# JSON OPERATIONS
# ============================================

@app.post("/format-json/")
async def format_json(json_text: str = Form(...)):
    """Format JSON with proper indentation"""
    json = lazy_import_json()
    
    try:
        parsed = json.loads(json_text)
        return {"formatted": json.dumps(parsed, indent=4)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")

# ============================================
# FILE CONVERSIONS
# ============================================

@app.post("/convert/pdf-to-docx")
async def convert_pdf_to_docx(file: UploadFile = File(...)):
    """Convert PDF to DOCX"""
    Converter = lazy_import_pdf2docx()
    
    input_path = await save_upload(file)
    output_path = input_path.replace(".pdf", ".docx")
    
    try:
        cv = Converter(input_path)
        cv.convert(output_path, start=0, end=None)
        cv.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
    return FileResponse(output_path, filename=os.path.basename(output_path))

@app.post("/convert/docx-to-pdf")
async def convert_docx_to_pdf(file: UploadFile = File(...)):
    """Convert DOCX to PDF"""
    Document = lazy_import_docx()
    canvas, A4 = lazy_import_reportlab()
    
    input_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    output_path = os.path.splitext(input_path)[0] + ".pdf"

    try:
        doc = Document(input_path)
        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        y = height - 50  # Start from top

        for para in doc.paragraphs:
            text = para.text
            if y < 50:  # Start new page
                c.showPage()
                y = height - 50
            c.drawString(50, y, text)
            y -= 15  # line spacing

        c.save()

        if not os.path.exists(output_path):
            raise Exception("Conversion failed: PDF not created")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DOCX → PDF conversion failed: {e}")

    return FileResponse(output_path, filename=os.path.basename(output_path))

@app.post("/convert/pdf-to-image")
async def convert_pdf_to_image(file: UploadFile = File(...), format: str = Form("jpg")):
    """Convert PDF to images"""
    convert_from_path = lazy_import_pdf2image()
    Image, _, _ = lazy_import_pil()
    zipfile = lazy_import_zipfile()
    
    try:
        # Save uploaded file
        temp_dir = tempfile.mkdtemp()
        input_path = os.path.join(temp_dir, file.filename)
        with open(input_path, "wb") as f:
            f.write(await file.read())

        # Convert PDF to images
        images = convert_from_path(input_path)
        output_files = []

        format_mapping = {"jpg": "JPEG", "png": "PNG"}
        for i, img in enumerate(images):
            img = img.convert("RGB")  # important for JPEG
            out_path = os.path.join(temp_dir, f"page_{i}.{format.lower()}")
            img.save(out_path, format_mapping.get(format.lower(), "JPEG"))
            output_files.append(out_path)

        # Single image -> return file, multiple -> zip
        if len(output_files) == 1:
            return FileResponse(
                output_files[0], 
                media_type=f"image/{format.lower()}", 
                filename=f"page_0.{format.lower()}"
            )
        else:
            zip_path = os.path.join(temp_dir, "images.zip")
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for f in output_files:
                    zipf.write(f, os.path.basename(f))
            return FileResponse(zip_path, media_type="application/zip", filename="images.zip")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {e}")

@app.post("/convert/image-to-pdf")
async def convert_image_to_pdf(file: UploadFile = File(...)):
    """Convert image to PDF"""
    Image, _, _ = lazy_import_pil()
    
    input_path = await save_upload(file)
    output_path = input_path + ".pdf"
    
    try:
        img = Image.open(input_path)
        img.convert("RGB").save(output_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
    return FileResponse(output_path, filename=os.path.basename(output_path))

@app.post("/convert/ppt-to-pdf")
async def convert_ppt_to_pdf(file: UploadFile = File(...)):
    """Convert PPT/PPTX to PDF"""
    slides = lazy_import_aspose_slides()
    
    input_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    try:
        # Load presentation
        pres = slides.Presentation(input_path)

        # Save to PDF
        output_path = os.path.splitext(input_path)[0] + ".pdf"
        pres.save(output_path, slides.export.SaveFormat.PDF)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PPT/PPTX → PDF conversion failed: {e}")

    return FileResponse(output_path, filename=os.path.basename(output_path))

@app.post("/convert/excel-to-pdf")
async def convert_excel_to_pdf(file: UploadFile = File(...)):
    """Convert Excel to PDF"""
    load_workbook = lazy_import_openpyxl()
    canvas, A4 = lazy_import_reportlab()
    
    input_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    output_path = os.path.splitext(input_path)[0] + ".pdf"

    try:
        wb = load_workbook(input_path)
        c = canvas.Canvas(output_path, pagesize=A4)
        width, height = A4
        margin = 50
        y = height - margin

        for sheet in wb.worksheets:
            c.setFont("Helvetica-Bold", 14)
            c.drawString(margin, y, f"Sheet: {sheet.title}")
            y -= 20
            c.setFont("Helvetica", 12)

            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                c.drawString(margin, y, row_text)
                y -= 15
                if y < margin:
                    c.showPage()
                    y = height - margin

            c.showPage()  # New page for next sheet

        c.save()

        if not os.path.exists(output_path):
            raise Exception("PDF not created")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel → PDF conversion failed: {e}")

    return FileResponse(output_path, filename=os.path.basename(output_path))
